import requests
import json
import pandas as pd
from openpyxl import Workbook

# URL base de la API
base_url = "https://www.tokkobroker.com/api"

# Endpoint específico y parámetros
endpoint = "/v1/property/"
params = {
    "lang": "es_ar",
    "format": "json",
    "limit": 1000,
    "key": "8c8e330d1213d990d7a1c5e568ecf82ebd240850"
}

# Hacer la solicitud GET a la API con los parámetros
response = requests.get(base_url + endpoint, params=params)

# Verificar si la solicitud fue exitosa (código 200)
if response.status_code == 200:
    # Convertir la respuesta JSON a un diccionario de Python
    data = response.json()
    
    # Especificar el nombre del archivo donde se guardará el JSON
    file_name = "Propiedades.json"
    
    # Guardar el diccionario de Python en un archivo JSON
    with open(file_name, 'w', encoding='utf-8') as json_file:
        json.dump(data, json_file, indent=4, ensure_ascii=False)
    
    print(f"Los datos JSON se han guardado correctamente en {file_name}")
else:
    print(f"Error en la solicitud: {response.status_code}")
    exit()  # Salir si la solicitud falla

# Cargar el JSON desde el archivo
json_path = r'C:\Users\renzo\Desktop\Trabajo Prone\Prone\Propiedades.json'
with open(json_path, 'r', encoding='utf-8') as json_file:
    data = json.load(json_file)

#Renzo C:\Users\renzo\Desktop\Trabajo Prone\Prone
# Crear un nuevo archivo Excel
wb = Workbook()
ws_property = wb.active
ws_property.title = "Propiedades"

# Definir todas las columnas posibles basadas en la estructura del JSON
header = [
    'id',"reference_code", 'address', 'age', 'bathroom_amount', 'created_at', 'description', 'expenses', 
    'fake_address', 'geo_lat', 'geo_long', 'is_starred_on_web', 'status', 'situation', 
    'type_id', 'type_name', 'operation_id', 'operation_type', 'price', 'currency', 
    'branch_id', 'branch_name', 'branch_email', 'branch_phone', 'branch_geo_lat', 
    'branch_geo_long', 'producer_id', 'producer_name', 'producer_email', 'location_id', 
    'location_name', 'location_full', 'publication_title', 'public_url'
]
ws_property.append(header)

# Función para truncar cadenas de texto a una longitud máxima
def truncate_string(value, max_length):
    return str(value)[:max_length] if value else ''

# Definir longitudes máximas
MAX_LENGTH_ID = 50
MAX_LENGTH_ADDRESS = 100
MAX_LENGTH_DESCRIPTION = 255

# Función para obtener el id o el campo más significativo de un objeto
def get_significant_field(item, field="id"):
    return item.get(field) if isinstance(item, dict) else None

# Recorrer los objetos y escribir cada fila en el Excel
for obj in data.get('objects', []):
    if obj.get('operations'):  # Verificar si hay operaciones disponibles
        operation = obj['operations'][0]  # Tomar la primera operación
        price_info = operation.get('prices', [{}])[0]  # Tomar el primer precio
    else:
        operation = {}
        price_info = {}

    # Escribir los datos en el Excel, truncando donde sea necesario
    row = [
        truncate_string(obj.get('id', ''), MAX_LENGTH_ID),
        truncate_string(obj.get('reference_code', ''), MAX_LENGTH_DESCRIPTION),
        truncate_string(obj.get('address', ''), MAX_LENGTH_ADDRESS),
        str(obj.get('age', '')),  # Convertir a cadena de texto
        str(obj.get('bathroom_amount', '')),  # Convertir a cadena de texto
        str(obj.get('created_at', '')),  # Dejar en formato de cadena de texto
        truncate_string(obj.get('description', ''), MAX_LENGTH_DESCRIPTION),
        str(obj.get('expenses', '')),  # Convertir a cadena de texto
        truncate_string(obj.get('fake_address', ''), MAX_LENGTH_ADDRESS),
        str(obj.get('geo_lat', '')),  # Convertir a cadena de texto
        str(obj.get('geo_long', '')),  # Convertir a cadena de texto
        str(obj.get('is_starred_on_web', '')),  # Convertir a cadena de texto
        str(obj.get('status', '')),  # Convertir a cadena de texto
        str(obj.get('situation', '')),  # Convertir a cadena de texto
        truncate_string(obj.get('type', {}).get('id', ''), MAX_LENGTH_ID),
        truncate_string(obj.get('type', {}).get('name', ''), MAX_LENGTH_DESCRIPTION),
        truncate_string(operation.get('operation_id', ''), MAX_LENGTH_ID),
        truncate_string(operation.get('operation_type', ''), MAX_LENGTH_DESCRIPTION),
        str(price_info.get('price', '')),  # Convertir a cadena de texto
        truncate_string(price_info.get('currency', ''), 3),  # Monedas suelen ser cortas
        truncate_string(obj.get('branch', {}).get('id', ''), MAX_LENGTH_ID),
        truncate_string(obj.get('branch', {}).get('name', ''), MAX_LENGTH_DESCRIPTION),
        truncate_string(obj.get('branch', {}).get('email', ''), MAX_LENGTH_DESCRIPTION),
        truncate_string(obj.get('branch', {}).get('phone', ''), MAX_LENGTH_DESCRIPTION),
        str(obj.get('branch', {}).get('geo_lat', '')),  # Convertir a cadena de texto
        str(obj.get('branch', {}).get('geo_long', '')),  # Convertir a cadena de texto
        truncate_string(obj.get('producer', {}).get('id', ''), MAX_LENGTH_ID),
        truncate_string(obj.get('producer', {}).get('name', ''), MAX_LENGTH_DESCRIPTION),
        truncate_string(obj.get('producer', {}).get('email', ''), MAX_LENGTH_DESCRIPTION),
        truncate_string(obj.get('location', {}).get('id', ''), MAX_LENGTH_ID),
        truncate_string(obj.get('location', {}).get('name', ''), MAX_LENGTH_DESCRIPTION),
        truncate_string(obj.get('location', {}).get('full_location', ''), MAX_LENGTH_DESCRIPTION),
        truncate_string(obj.get('publication_title', ''), MAX_LENGTH_DESCRIPTION),
        truncate_string(obj.get('public_url', ''), MAX_LENGTH_DESCRIPTION)
    ]
    ws_property.append(row)

# Crear hojas adicionales para listas detalladas
sheets = {
    "branch": ["id", "address", "name", "email", "phone", "geo_lat", "geo_long"],
    "location": ["id", "name", "full_location"],
    "producer": ["id", "name", "email", "cellphone"],
    "type": ["id", "name", "code"],
    "operations": ["operation_id", "property_id", "operation_type", "price", "currency"],
    "Property Owners": ['Owner ID', 'Owner Name', 'Owner Email', 'Owner Phone', 'Owner Work Email', 'Property ID']
}

# Crear hojas adicionales si no existen y agregar las cabeceras
for sheet_name, headers in sheets.items():
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)

# Llenar las hojas adicionales con los detalles de cada objeto
for obj in data.get('objects', []):
    # Branch
    branch_data = obj.get('branch', {})
    if branch_data:
        ws_branch = wb["branch"]
        ws_branch.append([
            branch_data.get('id', ''),
            branch_data.get('address', ''),
            branch_data.get('name', ''),
            branch_data.get('email', ''),
            branch_data.get('phone', ''),
            branch_data.get('geo_lat', ''),
            branch_data.get('geo_long', '')
        ])

    # Property Owners
    property_id = obj.get('id', '')  # Obtener el ID de la propiedad actual
    internal_data = obj.get('internal_data', {})
    property_owners = internal_data.get('property_owners', [])

    for owner in property_owners:
        # Extraer los detalles de cada propietario
        owner_id = owner.get('id', '')
        owner_name = owner.get('name', '')
        owner_email = owner.get('email', '')
        owner_phone = owner.get('cellphone', '')
        owner_work_email = owner.get('work_email', '')

        # Agregar los datos del propietario a la hoja "Property Owners"
        ws_property_owners = wb["Property Owners"]
        ws_property_owners.append([
            owner_id, owner_name, owner_email, owner_phone, owner_work_email, property_id
        ])

    # Location
    location_data = obj.get('location', {})
    if location_data:
        ws_location = wb["location"]
        ws_location.append([
            location_data.get('id', ''),
            location_data.get('name', ''),
            location_data.get('full_location', '')
        ])

    # Producer
    producer_data = obj.get('producer', {})
    if producer_data:
        ws_producer = wb["producer"]
        ws_producer.append([
            producer_data.get('id', ''),
            producer_data.get('name', ''),
            producer_data.get('email', ''),
            producer_data.get('cellphone', '')
        ])

    # Type
    type_data = obj.get('type', {})
    if type_data:
        ws_type = wb["type"]
        ws_type.append([
            type_data.get('id', ''),
            type_data.get('name', ''),
            type_data.get('code', '')
        ])

    # Operations
    for operation in obj.get('operations', []):
        for price_info in operation.get('prices', []):
            ws_operations = wb["operations"]
            ws_operations.append([
                obj.get('id', ''),
                price_info.get('price', ''),
            ])

# Guardar el archivo Excel
excel_path = "Propiedades.xlsx"
wb.save(excel_path)
print(f"El archivo Excel {excel_path} ha sido generado correctamente.")
